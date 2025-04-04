import openpyxl
from openpyxl.styles import Alignment
import shutil
from openpyxl.utils import column_index_from_string, get_column_letter
import os

def convert_column_input(column_input):
    """Convert column input (letter or number) to column number"""
    if isinstance(column_input, str):
        # Remove any spaces and convert to uppercase
        column_input = column_input.strip().upper()
        try:
            return column_index_from_string(column_input)
        except:
            raise ValueError(f"Invalid column input: {column_input}")
    else:
        try:
            return int(column_input)
        except:
            raise ValueError(f"Invalid column input: {column_input}")

def merge_cells(filepath, header_row, harm_col_num):
    """Merge cells in the Excel file based on column B values."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    # Find the maximum row and column
    max_row = ws.max_row
    max_col = ws.max_column
    
    # First, find merge ranges in column B
    merge_ranges = []
    current_value = None
    start_row = header_row + 1
    
    for row in range(header_row + 1, max_row + 1):
        cell_value = ws.cell(row=row, column=2).value
        
        if cell_value != current_value:
            if start_row and start_row < row - 1:
                merge_ranges.append((start_row, row - 1))
            current_value = cell_value
            start_row = row
    
    # Add the last range if necessary
    if start_row and start_row < max_row:
        merge_ranges.append((start_row, max_row))
    
    # Process each merge range independently
    for start_row, end_row in merge_ranges:
        # Merge column B
        if start_row < end_row:  # Only merge if there are multiple rows
            ws.merge_cells(f'B{start_row}:B{end_row}')
            merged_cell = ws.cell(row=start_row, column=2)
            merged_cell.alignment = Alignment(vertical='center')
        
        # For each column after B
        for col in range(3, max_col + 1):
            # Skip harm column and the next two columns only if harm_col_num is provided
            if harm_col_num is not None and (col == harm_col_num or col == harm_col_num + 1 or col == harm_col_num + 2):
                continue
                
            # Find sub-ranges within this range that have the same value
            sub_ranges = []
            sub_start = start_row
            current_value = ws.cell(row=start_row, column=col).value
            
            for row in range(start_row + 1, end_row + 1):
                cell_value = ws.cell(row=row, column=col).value
                
                if cell_value != current_value:
                    if sub_start < row - 1:  # Only add if there are multiple rows
                        sub_ranges.append((sub_start, row - 1))
                    sub_start = row
                    current_value = cell_value
            
            # Add the last sub-range if necessary
            if sub_start < end_row:
                sub_ranges.append((sub_start, end_row))
            
            # Merge each sub-range
            for sub_start, sub_end in sub_ranges:
                if sub_start < sub_end:  # Only merge if there are multiple rows
                    ws.merge_cells(
                        start_row=sub_start,
                        start_column=col,
                        end_row=sub_end,
                        end_column=col
                    )
                    merged_cell = ws.cell(row=sub_start, column=col)
                    merged_cell.alignment = Alignment(vertical='center')
    
    return wb

def format_general(file_path, header_row, harm_id_col=None):
    """Format Excel file with general formatting rules"""
    # Verify file is xlsx
    if not file_path.lower().endswith('.xlsx'):
        raise ValueError("Please select a valid .xlsx file")
    
    # Create output path
    output_path = file_path.rsplit('.', 1)[0] + '_formatted.xlsx'
    
    # Copy the original file to the new location before processing
    shutil.copy2(file_path, output_path)
    
    # Convert column input to number if harm_id_col is provided
    harm_col_num = None
    if harm_id_col is not None:
        try:
            harm_col_num = convert_column_input(harm_id_col)
        except ValueError:
            # If harm_id_col is invalid, just set it to None
            harm_col_num = None
    
    # Convert header_row to integer
    header_row = int(header_row)
    
    # Process the file
    return merge_cells(output_path, header_row, harm_col_num)