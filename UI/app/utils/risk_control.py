import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

def convert_column_input(column_input):
    """
    Convert Excel column input to column number.
    Handles both letter (A, B, AA, etc.) and number (1, 2, etc.) inputs.
    
    Examples:
        'A' -> 1
        'B' -> 2
        'Z' -> 26
        'AA' -> 27
        '1' -> 1
        1 -> 1
    
    Args:
        column_input: String or integer representing Excel column
    
    Returns:
        Integer representing column number
    
    Raises:
        ValueError: If input is invalid
    """
    if isinstance(column_input, str):
        # Remove any spaces and convert to uppercase
        column_input = column_input.strip().upper()
        
        # If it's a number in string form, convert it
        if column_input.isdigit():
            return int(column_input)
            
        # Handle letter input using openpyxl's utility
        try:
            return column_index_from_string(column_input)
        except ValueError:
            raise ValueError(f"Invalid column input: {column_input}. Please use letters (A, B, AA) or numbers.")
    
    # Handle direct number input
    if isinstance(column_input, (int, float)):
        try:
            column_num = int(column_input)
            if column_num <= 0:
                raise ValueError
            return column_num
        except ValueError:
            raise ValueError(f"Invalid column input: {column_input}. Column number must be positive.")
    
    raise ValueError(f"Invalid column input type: {type(column_input)}. Please use letters (A, B, AA) or numbers.")

def get_merged_cell_ranges(worksheet):
    """
    Get a dictionary of merged cell ranges and their values.
    Returns: Dict[tuple, str] where tuple is (start_row, end_row, column)
    """
    merged_ranges = {}
    
    # Get all merged ranges in the worksheet
    for merged_range in worksheet.merged_cells.ranges:
        # Only process single-column merged ranges
        if merged_range.min_col == merged_range.max_col:
            col = merged_range.min_col
            start_row = merged_range.min_row
            end_row = merged_range.max_row
            # Get value from the top cell of the merged range
            value = worksheet.cell(row=start_row, column=col).value
            merged_ranges[(start_row, end_row, col)] = value
            
    return merged_ranges

def merge_risk_control(risk_file, control_file, 
                      risk_header_row, control_header_row,
                      risk_id_col, control_id_col,
                      paste_col, control_content_col):
    """
    Merge risk control information from control document into risk document.
    Handles merged cells and maintains merged cell structure.
    """
    
    # Convert column inputs to numbers
    risk_id_col_num = convert_column_input(risk_id_col)
    control_id_col_num = convert_column_input(control_id_col)
    paste_col_num = convert_column_input(paste_col)
    control_content_col_num = convert_column_input(control_content_col)
    
    # Load workbooks
    risk_wb = openpyxl.load_workbook(risk_file)
    control_wb = openpyxl.load_workbook(control_file)
    
    risk_ws = risk_wb.active
    control_ws = control_wb.active
    
    # Get merged ranges for both worksheets
    risk_merged = get_merged_cell_ranges(risk_ws)
    control_merged = get_merged_cell_ranges(control_ws)
    
    # Create dictionary to store control content
    control_dict = {}
    
    # Process control document
    for row in range(control_header_row + 1, control_ws.max_row + 1):
        control_id = None
        control_content = None
        
        # Check if current row is part of a merged range in ID column
        for (start_row, end_row, col) in control_merged:
            if col == control_id_col_num and start_row <= row <= end_row:
                control_id = control_merged[(start_row, end_row, col)]
                break
        
        # If not in merged range, get direct cell value
        if control_id is None:
            control_id = control_ws.cell(row=row, column=control_id_col_num).value
        
        # Similarly for content column
        for (start_row, end_row, col) in control_merged:
            if col == control_content_col_num and start_row <= row <= end_row:
                control_content = control_merged[(start_row, end_row, col)]
                break
                
        if control_content is None:
            control_content = control_ws.cell(row=row, column=control_content_col_num).value
            
        # Store in dictionary, handling multiple controls for same ID
        if control_id and control_content:
            if control_id in control_dict:
                if isinstance(control_dict[control_id], list):
                    control_dict[control_id].append(control_content)
                else:
                    control_dict[control_id] = [control_dict[control_id], control_content]
            else:
                control_dict[control_id] = control_content
    
    # Process risk document
    current_merge_range = None
    
    for row in range(risk_header_row + 1, risk_ws.max_row + 1):
        risk_id = None
        
        # Check if current row is part of a merged range in ID column
        for (start_row, end_row, col) in risk_merged:
            if col == risk_id_col_num and start_row <= row <= end_row:
                risk_id = risk_merged[(start_row, end_row, col)]
                # Store merge range for paste column
                current_merge_range = (start_row, end_row)
                break
        
        # If not in merged range, get direct cell value
        if risk_id is None:
            risk_id = risk_ws.cell(row=row, column=risk_id_col_num).value
            current_merge_range = None
        
        # Handle content pasting and merging
        if current_merge_range and row == current_merge_range[0]:  # First row of merge range
            start_row, end_row = current_merge_range
            
            # Get content if we have a matching control
            content = None
            if risk_id and risk_id in control_dict:
                content = control_dict[risk_id]
                if isinstance(content, list):
                    content = "\n".join(str(item) for item in content if item is not None)
            
            # Write content (or empty string) and merge cells
            risk_ws.cell(row=start_row, column=paste_col_num).value = content
            
            # Always merge cells if range spans multiple rows, even if content is None
            if start_row != end_row:
                risk_ws.merge_cells(
                    start_row=start_row,
                    start_column=paste_col_num,
                    end_row=end_row,
                    end_column=paste_col_num
                )
        
        # Handle non-merged cells
        elif not current_merge_range:
            content = None
            if risk_id and risk_id in control_dict:
                content = control_dict[risk_id]
                if isinstance(content, list):
                    content = "\n".join(str(item) for item in content if item is not None)
            risk_ws.cell(row=row, column=paste_col_num).value = content
    
    return risk_wb