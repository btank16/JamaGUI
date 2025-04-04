import openpyxl
from openpyxl.styles import Alignment
import shutil
from openpyxl.utils import column_index_from_string, get_column_letter
import os

def convert_column_input(column_input):
    """Convert column input (letter or number) to column number"""
    if isinstance(column_input, str):
        # Remove any spaces and convert to uppercase
        column_input = column_input.strip().upper()
        try:
            return column_index_from_string(column_input)
        except:
            raise ValueError(f"Invalid column input: {column_input}")
    else:
        try:
            return int(column_input)
        except:
            raise ValueError(f"Invalid column input: {column_input}")

def merge_cells(filepath, header_row, first_parent_col, second_parent_col, harm_col_num):
    """Merge cells in the Excel file based on two parent columns."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    max_row = ws.max_row
    max_col = ws.max_column
    
    # First, find merge ranges for the first parent column
    first_parent_ranges = []
    current_value = None
    start_row = header_row + 1
    
    for row in range(header_row + 1, max_row + 1):
        cell_value = ws.cell(row=row, column=first_parent_col).value
        
        if cell_value != current_value:
            if start_row and start_row < row - 1:
                first_parent_ranges.append((start_row, row - 1))
            current_value = cell_value
            start_row = row
    
    # Add the last range if necessary
    if start_row and start_row < max_row:
        first_parent_ranges.append((start_row, max_row))
    
    # Process each first parent range
    for parent_start, parent_end in first_parent_ranges:
        # Merge first parent column
        if parent_start < parent_end:
            ws.merge_cells(
                start_row=parent_start,
                start_column=first_parent_col,
                end_row=parent_end,
                end_column=first_parent_col
            )
            merged_cell = ws.cell(row=parent_start, column=first_parent_col)
            merged_cell.alignment = Alignment(vertical='center')
        
        # Find and merge ranges for columns between first and second parent
        for col in range(first_parent_col + 1, second_parent_col):
            if harm_col_num is not None and (col == harm_col_num or col == harm_col_num + 1 or col == harm_col_num + 2):
                continue
                
            current_value = None
            sub_start = parent_start
            
            for row in range(parent_start, parent_end + 1):
                cell_value = ws.cell(row=row, column=col).value
                
                if cell_value != current_value:
                    if sub_start < row - 1:
                        ws.merge_cells(
                            start_row=sub_start,
                            start_column=col,
                            end_row=row - 1,
                            end_column=col
                        )
                        merged_cell = ws.cell(row=sub_start, column=col)
                        merged_cell.alignment = Alignment(vertical='center')
                    current_value = cell_value
                    sub_start = row
            
            if sub_start < parent_end:
                ws.merge_cells(
                    start_row=sub_start,
                    start_column=col,
                    end_row=parent_end,
                    end_column=col
                )
                merged_cell = ws.cell(row=sub_start, column=col)
                merged_cell.alignment = Alignment(vertical='center')
        
        # Process second parent ranges within first parent range
        second_parent_ranges = []
        current_value = None
        sub_start = parent_start
        
        for row in range(parent_start, parent_end + 1):
            cell_value = ws.cell(row=row, column=second_parent_col).value
            
            if cell_value != current_value:
                if sub_start < row - 1:
                    second_parent_ranges.append((sub_start, row - 1))
                current_value = cell_value
                sub_start = row
        
        if sub_start < parent_end:
            second_parent_ranges.append((sub_start, parent_end))
        
        # Process each second parent range
        for sub_start, sub_end in second_parent_ranges:
            # Merge second parent column
            if sub_start < sub_end:
                ws.merge_cells(
                    start_row=sub_start,
                    start_column=second_parent_col,
                    end_row=sub_end,
                    end_column=second_parent_col
                )
                merged_cell = ws.cell(row=sub_start, column=second_parent_col)
                merged_cell.alignment = Alignment(vertical='center')
            
            # Merge remaining columns within second parent bounds
            for col in range(second_parent_col + 1, max_col + 1):
                if harm_col_num is not None and (col == harm_col_num or col == harm_col_num + 1 or col == harm_col_num + 2):
                    continue
                    
                current_value = None
                merge_start = sub_start
                
                for row in range(sub_start, sub_end + 1):
                    cell_value = ws.cell(row=row, column=col).value
                    
                    if cell_value != current_value:
                        if merge_start < row - 1:
                            ws.merge_cells(
                                start_row=merge_start,
                                start_column=col,
                                end_row=row - 1,
                                end_column=col
                            )
                            merged_cell = ws.cell(row=merge_start, column=col)
                            merged_cell.alignment = Alignment(vertical='center')
                        current_value = cell_value
                        merge_start = row
                
                if merge_start < sub_end:
                    ws.merge_cells(
                        start_row=merge_start,
                        start_column=col,
                        end_row=sub_end,
                        end_column=col
                    )
                    merged_cell = ws.cell(row=merge_start, column=col)
                    merged_cell.alignment = Alignment(vertical='center')
    
    return wb

def format_urra(file_path, header_row, first_parent_col, second_parent_col, harm_id_col=None):
    """Format Excel file with URRA-specific formatting rules"""
    if not file_path.lower().endswith('.xlsx'):
        raise ValueError("Please select a valid .xlsx file")
    
    output_path = file_path.rsplit('.', 1)[0] + '_formatted.xlsx'
    shutil.copy2(file_path, output_path)
    
    # Convert column inputs to numbers
    first_parent_num = convert_column_input(first_parent_col)
    second_parent_num = convert_column_input(second_parent_col)
    harm_col_num = None if harm_id_col is None else convert_column_input(harm_id_col)
    
    # Validate column order
    if second_parent_num <= first_parent_num:
        raise ValueError("Second parent column must be after first parent column")
    
    header_row = int(header_row)
    
    return merge_cells(output_path, header_row, first_parent_num, second_parent_num, harm_col_num)