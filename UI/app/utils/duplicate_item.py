from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string

def get_merged_cell_value(worksheet, row, col):
    """Get the value of a cell, taking into account if it's part of a merged range."""
    for merged_range in worksheet.merged_cells.ranges:
        if merged_range.min_row <= row <= merged_range.max_row and \
           merged_range.min_col <= col <= merged_range.max_col:
            return worksheet.cell(merged_range.min_row, merged_range.min_col).value
    return worksheet.cell(row, col).value

def highlight_duplicates_in_column(file_path: str, column_letter: str) -> None:
    """
    Creates a new worksheet with simplified content from the specified column and 
    highlights duplicates in the new worksheet. Also includes content from the previous column.
    
    Args:
        file_path (str): Path to the Excel file
        column_letter (str): Column letter to search (e.g., 'A', 'B', 'C')
    """
    # Load the workbook and select active sheet
    wb = load_workbook(file_path)
    source_ws = wb.active
    
    # Create a new worksheet for simplified content
    analysis_ws = wb.create_sheet("Duplicate_Analysis")
    
    # Convert column letter to index and get previous column
    col_idx = column_index_from_string(column_letter)
    prev_col_idx = max(1, col_idx - 1)  # Ensure we don't go below column A
    prev_col_letter = get_column_letter(prev_col_idx)
    
    # Add headers to analysis worksheet
    analysis_ws['A1'] = f"Content from Column {prev_col_letter}"
    analysis_ws['B1'] = f"Content from Column {column_letter}"
    analysis_ws['C1'] = "Original Row(s)"
    analysis_ws['D1'] = "Item ID"
    
    # Process each row in the source worksheet
    processed_rows = set()
    analysis_row = 2  # Start after header
    
    for row in range(1, source_ws.max_row + 1):
        if row in processed_rows:
            continue
            
        value = get_merged_cell_value(source_ws, row, col_idx)
        prev_value = get_merged_cell_value(source_ws, row, prev_col_idx)
        
        if value:  # Skip empty cells
            value = str(value).strip()
            prev_value = str(prev_value).strip() if prev_value else ""
            
            # Check if this cell is part of a merged range
            is_merged = False
            merged_range = None
            for range_ in source_ws.merged_cells.ranges:
                if (range_.min_col <= col_idx <= range_.max_col and
                    range_.min_row <= row <= range_.max_row):
                    is_merged = True
                    merged_range = range_
                    break
            
            if is_merged:
                # Add all rows in merge range to processed set
                for r in range(merged_range.min_row, merged_range.max_row + 1):
                    processed_rows.add(r)
                # Store original row range information
                row_info = f"Rows {merged_range.min_row}-{merged_range.max_row}"
            else:
                row_info = f"Row {row}"
            
            # Write to analysis worksheet
            analysis_ws.cell(row=analysis_row, column=1, value=prev_value)
            analysis_ws.cell(row=analysis_row, column=2, value=value)
            analysis_ws.cell(row=analysis_row, column=3, value=row_info)
            analysis_row += 1
    
    # Create yellow fill pattern
    yellow_fill = PatternFill(start_color='FFFF00',
                             end_color='FFFF00',
                             fill_type='solid')
    
    # Find and highlight duplicates in analysis worksheet
    value_positions = {}
    for row in range(2, analysis_ws.max_row + 1):  # Start after header
        value = analysis_ws.cell(row=row, column=2).value  # Check column B for duplicates
        if value:
            value = str(value).strip()
            if value in value_positions:
                value_positions[value].append(row)
            else:
                value_positions[value] = [row]
    
    # Highlight duplicates and assign IDs
    current_id = 1
    processed_values = set()
    
    for value, rows in value_positions.items():
        if value not in processed_values:
            processed_values.add(value)
            # Assign the same ID to all instances of this value
            for row in rows:
                analysis_ws.cell(row=row, column=4, value=current_id)  # Add ID
                if len(rows) > 1:  # If there are duplicates
                    analysis_ws.cell(row=row, column=1).fill = yellow_fill
                    analysis_ws.cell(row=row, column=2).fill = yellow_fill
                    analysis_ws.cell(row=row, column=3).fill = yellow_fill
                    analysis_ws.cell(row=row, column=4).fill = yellow_fill
            current_id += 1
    
    # Adjust column widths for better readability
    analysis_ws.column_dimensions['A'].width = 50
    analysis_ws.column_dimensions['B'].width = 50
    analysis_ws.column_dimensions['C'].width = 20
    analysis_ws.column_dimensions['D'].width = 10  # Add width for new column
    
    # Save the workbook
    wb.save(file_path)
