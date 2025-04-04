import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string

def convert_column_input(column_input):
    """Convert column input (letter or number) to column number"""
    if isinstance(column_input, str):
        # Remove any spaces and convert to uppercase
        column_input = column_input.strip().upper()
        try:
            return column_index_from_string(column_input)
        except:
            raise ValueError(f"Invalid column input: {column_input}")
    else:
        try:
            return int(column_input)
        except:
            raise ValueError(f"Invalid column input: {column_input}")

def extract_number(value):
    if isinstance(value, (int, float)):
        return value
    elif isinstance(value, str):
        return int(''.join(filter(str.isdigit, value))) if any(char.isdigit() for char in value) else None
    return None

def determine_output(parent_value, child_value, risk_matrix_values=None):
    # If risk_matrix_values is provided, use it
    if risk_matrix_values:
        parent_idx = 5 - parent_value  # Convert 1-5 to array index
        child_idx = child_value - 1    # Convert 1-5 to array index
        
        if 0 <= parent_idx < 5 and 0 <= child_idx < 5:
            risk_level = risk_matrix_values[parent_idx][child_idx]
            color = None
            if risk_level == "INT":
                color = "red"
            elif risk_level == "MOD":
                color = "orange"
            return risk_level, color
    # Otherwise use default logic
    else:
        if parent_value == 5:
            if child_value == 1:
                return "LOW", None
            elif child_value in [2, 3]:
                return "MOD", "orange"
            elif child_value in [4, 5]:
                return "INT", "red"
        elif parent_value == 4:
            if child_value == 1:
                return "LOW", None
            elif child_value in [2, 3, 4]:
                return "MOD", "orange"
            elif child_value == 5:
                return "INT", "red"
        elif parent_value == 3:
            if child_value in [1, 2, 3]:
                return "LOW", None
            elif child_value == 4:
                return "MOD", "orange"
            elif child_value == 5:
                return "INT", "red"
        elif parent_value == 2:
            if child_value in [1, 2, 3, 4]:
                return "LOW", None
            elif child_value == 5:
                return "MOD", "orange"
        elif parent_value == 1:
            return "LOW", None
    return None, None

def calculate_risk_score(wb, occurrence_col, severity_col, risk_analysis_col, header_row, risk_matrix_values=None):
    """Calculate risk scores based on occurrence and severity"""
    ws = wb.active
    
    # Convert inputs to column numbers
    header_row = int(header_row)
    occurrence_col = convert_column_input(occurrence_col)
    severity_col = convert_column_input(severity_col)
    risk_analysis_col = convert_column_input(risk_analysis_col)
    
    # Create fill patterns
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Start processing from after header row
    row = header_row + 1
    while row <= ws.max_row:
        parent_cell = ws.cell(row=row, column=occurrence_col)
        parent_value = extract_number(parent_cell.value)
        
        # Check if the cell is part of a merged range
        if parent_cell.coordinate in ws.merged_cells:
            merged_range = [cell_range for cell_range in ws.merged_cells.ranges 
                          if parent_cell.coordinate in cell_range][0]
            end_row = merged_range.max_row
        else:
            end_row = row

        # Process all child rows corresponding to the parent
        for child_row in range(row, end_row + 1):
            child_value = extract_number(ws.cell(row=child_row, column=severity_col).value)
            
            if parent_value is not None and child_value is not None:
                output_text, fill_color = determine_output(parent_value, child_value, risk_matrix_values)
                
                if output_text:
                    output_cell = ws.cell(row=child_row, column=risk_analysis_col)
                    output_cell.value = output_text
                    if fill_color == "orange":
                        output_cell.fill = orange_fill
                    elif fill_color == "red":
                        output_cell.fill = red_fill

        row = end_row + 1
    
    return wb 