from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter

def add_merged_borders(wb):
    """Add thick bottom borders after merged cell groups in column B"""
    ws = wb.active
    
    # Find the last column with content
    max_col = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and cell.column > max_col:
                max_col = cell.column
    
    # Create thick bottom border style
    thick_border = Border(bottom=Side(style='thick'))
    
    # Track the last row of each merged range in column B
    merge_end_rows = set()
    for merged_range in ws.merged_cells.ranges:
        # Check if this merge range includes column B
        if merged_range.min_col <= 2 <= merged_range.max_col:
            merge_end_rows.add(merged_range.max_row)
    
    # Add borders to each row that ends a merge group
    for row_num in merge_end_rows:
        for col in range(2, max_col + 1):  # Start from B column
            cell = ws.cell(row=row_num, column=col)
            current_border = cell.border
            
            # Create new border maintaining existing styles
            new_border = Border(
                left=current_border.left,
                right=current_border.right,
                top=current_border.top,
                bottom=Side(style='thick')
            )
            cell.border = new_border
    
    return wb 