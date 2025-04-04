import openpyxl
from openpyxl.styles import Alignment
import shutil
from .column_converter import convert_column_input

def usertask_format(file_path, header_row, item_type_col, us_name_col, ut_name_col):
    """Format Excel file with Marathon UT specific formatting rules"""
    if not file_path.lower().endswith('.xlsx'):
        raise ValueError("Please select a valid .xlsx file")
    
    output_path = file_path.rsplit('.', 1)[0] + '_formatted.xlsx'
    shutil.copy2(file_path, output_path)
    
    # Convert column inputs to numbers
    item_type_col_num = convert_column_input(item_type_col)
    us_name_col_num = convert_column_input(us_name_col)
    ut_name_col_num = convert_column_input(ut_name_col)
    
    # Convert header_row to integer
    header_row = int(header_row)
    
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    
    # Insert new column to the left of Item Type column
    ws.insert_cols(item_type_col_num)
    
    # First pass: Just populate the new column without merging
    folder_positions = []  # Track where folders are located
    for row in range(header_row + 1, ws.max_row + 1):
        item_type = ws.cell(row=row, column=item_type_col_num + 1).value
        next_item_type = ws.cell(row=row + 1, column=item_type_col_num + 1).value if row < ws.max_row else None
        
        if item_type == "Folder":
            folder_positions.append(row)
            current_folder = ws.cell(row=row, column=us_name_col_num + 1).value
            paste_row = row + 1 if next_item_type != "Folder" else row
            ws.cell(row=paste_row, column=item_type_col_num).value = current_folder
    
    # Save the workbook after column population
    wb.save(output_path)
    
    # Second pass: Handle merging
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    
    # Process merging for the new column
    merge_start = None
    current_value = None
    
    for row in range(header_row + 1, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=item_type_col_num).value
        
        if cell_value and cell_value != current_value:
            if merge_start and current_value:
                # Find the next folder position
                next_folder = next((pos for pos in folder_positions if pos > merge_start), ws.max_row + 1)
                end_row = next_folder - 1
                
                ws.merge_cells(
                    start_row=merge_start,
                    start_column=item_type_col_num,
                    end_row=end_row,
                    end_column=item_type_col_num
                )
                merged_cell = ws.cell(row=merge_start, column=item_type_col_num)
                merged_cell.alignment = Alignment(vertical='center')
            merge_start = row
            current_value = cell_value
    
    # Handle the last merge for the new column
    if merge_start and current_value:
        ws.merge_cells(
            start_row=merge_start,
            start_column=item_type_col_num,
            end_row=ws.max_row,
            end_column=item_type_col_num
        )
        merged_cell = ws.cell(row=merge_start, column=item_type_col_num)
        merged_cell.alignment = Alignment(vertical='center')
    
    # Process parent column (US Name) merging
    merge_start = None
    current_value = None
    parent_ranges = []  # Store parent merge ranges
    
    for row in range(header_row + 1, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=us_name_col_num + 1).value
        
        if cell_value != current_value:
            if merge_start and current_value is not None and merge_start < row - 1:
                ws.merge_cells(
                    start_row=merge_start,
                    start_column=us_name_col_num + 1,
                    end_row=row - 1,
                    end_column=us_name_col_num + 1
                )
                parent_ranges.append((merge_start, row - 1))
                merged_cell = ws.cell(row=merge_start, column=us_name_col_num + 1)
                merged_cell.alignment = Alignment(vertical='center')
            merge_start = row
            current_value = cell_value
    
    # Handle last parent merge
    if merge_start and current_value is not None and merge_start < ws.max_row:
        ws.merge_cells(
            start_row=merge_start,
            start_column=us_name_col_num + 1,
            end_row=ws.max_row,
            end_column=us_name_col_num + 1
        )
        parent_ranges.append((merge_start, ws.max_row))
        merged_cell = ws.cell(row=merge_start, column=us_name_col_num + 1)
        merged_cell.alignment = Alignment(vertical='center')
    
    # Process columns between parent and child within parent bounds
    for col in range(us_name_col_num + 2, ut_name_col_num + 1):
        for start_row, end_row in parent_ranges:
            current_value = ws.cell(row=start_row, column=col).value
            if current_value is not None:
                ws.merge_cells(
                    start_row=start_row,
                    start_column=col,
                    end_row=end_row,
                    end_column=col
                )
                merged_cell = ws.cell(row=start_row, column=col)
                merged_cell.alignment = Alignment(vertical='center')
    
    # First process child column and store its merge ranges
    child_ranges = []
    merge_start = None
    current_value = None
    
    for row in range(header_row + 1, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=ut_name_col_num + 1).value
        
        if cell_value != current_value:
            if merge_start and current_value is not None and merge_start < row - 1:
                ws.merge_cells(
                    start_row=merge_start,
                    start_column=ut_name_col_num + 1,
                    end_row=row - 1,
                    end_column=ut_name_col_num + 1
                )
                child_ranges.append((merge_start, row - 1, current_value))
                merged_cell = ws.cell(row=merge_start, column=ut_name_col_num + 1)
                merged_cell.alignment = Alignment(vertical='center')
            merge_start = row
            current_value = cell_value
    
    # Handle last child merge
    if merge_start and current_value is not None and merge_start < ws.max_row:
        ws.merge_cells(
            start_row=merge_start,
            start_column=ut_name_col_num + 1,
            end_row=ws.max_row,
            end_column=ut_name_col_num + 1
        )
        child_ranges.append((merge_start, ws.max_row, current_value))
        merged_cell = ws.cell(row=merge_start, column=ut_name_col_num + 1)
        merged_cell.alignment = Alignment(vertical='center')
    
    # Process following columns based on child ranges
    for col in range(ut_name_col_num + 2, ws.max_column + 1):
        for start_row, end_row, _ in child_ranges:
            cell_value = ws.cell(row=start_row, column=col).value
            if cell_value is not None:
                ws.merge_cells(
                    start_row=start_row,
                    start_column=col,
                    end_row=end_row,
                    end_column=col
                )
                merged_cell = ws.cell(row=start_row, column=col)
                merged_cell.alignment = Alignment(vertical='center')
    
    wb.save(output_path)
    return wb 