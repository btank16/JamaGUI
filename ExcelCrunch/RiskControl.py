#This script can combine the risk controls traced in Jama to the dFMEA table
#Export and format both the dFMEA and trace to risk controls as separate excel files
#The first file you select is the dFMEA, and the second is the risk controls
#Note if the first cell in the last column is empty, the script will overwrite this column (add placeholder text if needed)

import openpyxl
import tkinter as tk
from tkinter import filedialog

def select_file():
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    return file_path

def find_first_empty_column(sheet):
    for col in range(1, sheet.max_column + 2):
        if all(sheet.cell(row=1, column=col).value is None for row in range(1, sheet.max_row + 1)):
            return col

def get_cell_ranges(sheet):
    cell_ranges = {}
    current_value = None
    start_row = 1

    for row in range(1, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=1).value
        if cell_value is not None:
            if current_value is not None:
                cell_ranges[current_value] = (start_row, row - 1)
            current_value = cell_value
            start_row = row

    if current_value is not None:
        cell_ranges[current_value] = (start_row, sheet.max_row)

    return cell_ranges

def combine_excel_sheets():
    print("Select the Excel file with the dFMEA:")
    file1 = select_file()
    print("Select the Excel file with the Risk Controls:")
    file2 = select_file()

    # Load workbooks
    wb1 = openpyxl.load_workbook(file1)
    wb2 = openpyxl.load_workbook(file2)
    sheet1 = wb1.active
    sheet2 = wb2.active

    # Find the first empty column in the first file
    empty_col = find_first_empty_column(sheet1)

    # Get cell ranges for both sheets
    ranges1 = get_cell_ranges(sheet1)
    ranges2 = get_cell_ranges(sheet2)

    # Process each cell range in the first file
    for id1, (start_row1, end_row1) in ranges1.items():
        if id1 in ranges2:
            start_row2, end_row2 = ranges2[id1]
            
            # Collect content from column B in the second file
            content = []
            for row in range(start_row2, end_row2 + 1):
                cell_value = sheet2.cell(row=row, column=2).value
                if cell_value is not None:
                    content.append(str(cell_value))
            
            # Determine the content to write
            if content:
                merged_content = "\n".join(content)
            else:
                merged_content = "No risk control applied"
            
            # Merge cells in the first file if the range spans multiple rows
            if start_row1 != end_row1:
                sheet1.merge_cells(start_row=start_row1, start_column=empty_col, 
                                   end_row=end_row1, end_column=empty_col)
            merged_cell = sheet1.cell(row=start_row1, column=empty_col)
            merged_cell.value = merged_content

    # Save the updated first workbook
    output_file = "updated_excel.xlsx"
    wb1.save(output_file)
    print(f"Updated Excel file saved as {output_file}")

if __name__ == "__main__":
    combine_excel_sheets()