# This sheet calculates the risk analysis in an excel sheet
#The first column you select is the Occurrence
#The second column you select is the Severity
#The third column you select is the Risk Analysis
#If you need to do Pre- and Post- risk analysis, run the script twice

import openpyxl
from tkinter import Tk, filedialog
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

def select_file():
    root = Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    return file_path

def extract_number(value):
    if isinstance(value, (int, float)):
        return value
    elif isinstance(value, str):
        return int(''.join(filter(str.isdigit, value))) if any(char.isdigit() for char in value) else None
    return None

def determine_output(parent_value, child_value):
    if parent_value == 5:
        if child_value == 1:
            return "LOW", None
        elif child_value in [2, 3]:
            return "MOD", "orange"
        elif child_value in [4, 5]:
            return "INT", "red"
    elif parent_value == 4:
        if child_value == 1:
            return "LOW", None
        elif child_value in [2, 3, 4]:
            return "MOD", "orange"
        elif child_value == 5:
            return "INT", "red"
    elif parent_value == 3:
        if child_value in [1, 2, 3]:
            return "LOW", None
        elif child_value == 4:
            return "MOD", "orange"
        elif child_value == 5:
            return "INT", "red"
    elif parent_value == 2:
        if child_value in [1, 2, 3, 4]:
            return "LOW", None
        elif child_value == 5:
            return "MOD", "orange"
    elif parent_value == 1:
        return "LOW", None
    return None, None

def process_excel(input_file, parent_col, child_col, output_col):
    workbook = openpyxl.load_workbook(input_file)
    sheet = workbook.active

    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Process the data
    row = 1
    while row <= sheet.max_row:
        parent_cell = sheet.cell(row=row, column=parent_col)
        parent_value = extract_number(parent_cell.value)
        
        # Check if the cell is part of a merged range
        if parent_cell.coordinate in sheet.merged_cells:
            merged_range = [cell_range for cell_range in sheet.merged_cells.ranges if parent_cell.coordinate in cell_range][0]
            end_row = merged_range.max_row
        else:
            end_row = row

        # Process all child rows corresponding to the parent
        for child_row in range(row, end_row + 1):
            child_value = extract_number(sheet.cell(row=child_row, column=child_col).value)
            
            if parent_value is not None and child_value is not None:
                output_text, fill_color = determine_output(parent_value, child_value)
                
                if output_text:
                    output_cell = sheet.cell(row=child_row, column=output_col)
                    output_cell.value = output_text
                    if fill_color == "orange":
                        output_cell.fill = orange_fill
                    elif fill_color == "red":
                        output_cell.fill = red_fill

        row = end_row + 1

    # Save the modified workbook
    output_file = input_file.rsplit('.', 1)[0] + '_processed.xlsx'
    workbook.save(output_file)
    print(f"Processed file saved as: {output_file}")

def main():
    input_file = select_file()
    if not input_file:
        print("No file selected. Exiting.")
        return

    parent_col = int(input("Enter the column number for Occurrence (e.g., 1 for column A): "))
    child_col = int(input("Enter the column number for Severity (e.g., 2 for column B): "))
    output_col = int(input("Enter the column number for Risk Analysis (e.g., 3 for column C): "))

    process_excel(input_file, parent_col, child_col, output_col)

if __name__ == "__main__":
    main()