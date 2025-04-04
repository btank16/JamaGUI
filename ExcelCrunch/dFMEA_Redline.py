import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter, column_index_from_string
import tkinter as tk
from tkinter import filedialog
import warnings

# Suppress UserWarnings
warnings.filterwarnings("ignore", category=UserWarning)

def get_merged_cell_value(worksheet, row, col):
    """Get the value of a cell, taking into account if it's part of a merged range."""
    if isinstance(col, str):
        col = column_index_from_string(col)
    
    for merged_range in worksheet.merged_cells.ranges:
        if merged_range.min_row <= row <= merged_range.max_row and \
           merged_range.min_col <= col <= merged_range.max_col:
            return worksheet.cell(merged_range.min_row, merged_range.min_col).value
    return worksheet.cell(row, col).value

def get_merged_cell_range(worksheet, row, col):
    """Get the range of a merged cell if the cell is part of a merged range."""
    if isinstance(col, str):
        col = column_index_from_string(col)
    
    for merged_range in worksheet.merged_cells.ranges:
        if merged_range.min_row <= row <= merged_range.max_row and \
           merged_range.min_col <= col <= merged_range.max_col:
            return merged_range
    return None

def find_matching_row(worksheet, value, column):
    """Find the row in the given worksheet where the value in the specified column matches."""
    if isinstance(column, int):
        column = get_column_letter(column)
    
    for row in range(1, worksheet.max_row + 1):
        if get_merged_cell_value(worksheet, row, column) == value:
            return row
    return None

def get_ids_in_range(worksheet, start_row, end_row, col):
    """Get all IDs within a range for a specific column, handling merged cells."""
    ids = []
    row = start_row
    
    # Convert column to integer if it's a string
    if isinstance(col, str):
        col = column_index_from_string(col)
    
    while row <= end_row:
        merged_range = get_merged_cell_range(worksheet, row, col)
        if merged_range:
            value = get_merged_cell_value(worksheet, row, col)
            if value is not None:
                ids.append(value)
            row = merged_range.max_row + 1
        else:
            value = worksheet.cell(row, col).value
            if value is not None:
                ids.append(value)
            row += 1
    return ids

def write_to_cell(worksheet, row, col, value, font=None):
    """Write to a cell, handling merged cells by writing to the top-left cell."""
    cell = worksheet.cell(row, col)
    cell.value = value
    if font:
        cell.font = font

def compare_excel_files(original_file, new_file, output_file):
    wb_original = openpyxl.load_workbook(original_file)
    wb_new = openpyxl.load_workbook(new_file)
    wb_output = openpyxl.Workbook()
    
    for sheet_name in wb_new.sheetnames:
        if sheet_name in wb_original.sheetnames:
            ws_original = wb_original[sheet_name]
            ws_new = wb_new[sheet_name]
            ws_output = wb_output.create_sheet(sheet_name)
            
            output_row = 1
            merge_ranges_to_apply = []
            
            new_row = 1
            while new_row <= ws_new.max_row:
                key_value = get_merged_cell_value(ws_new, new_row, 'A')
                original_row = find_matching_row(ws_original, key_value, 'A')
                
                # Write the key value to the output
                write_to_cell(ws_output, output_row, 1, key_value)
                
                new_merged_range_A = get_merged_cell_range(ws_new, new_row, 'A')
                if new_merged_range_A:
                    row_span_A = new_merged_range_A.max_row - new_merged_range_A.min_row + 1
                else:
                    row_span_A = 1
                
                start_output_row = output_row
                
                if original_row is not None:
                    original_merged_range_A = get_merged_cell_range(ws_original, original_row, 'A')
                    original_row_span_A = original_merged_range_A.max_row - original_merged_range_A.min_row + 1 if original_merged_range_A else 1
                    
                    # Compare columns B to F and M to N
                    for col in ['B', 'C', 'D', 'E', 'F', 'M', 'N']:
                        new_ids = get_ids_in_range(ws_new, new_row, new_row + row_span_A - 1, col)
                        original_ids = get_ids_in_range(ws_original, original_row, original_row + original_row_span_A - 1, col)
                        
                        all_ids = list(set(new_ids + original_ids))
                        for i, id_value in enumerate(all_ids):
                            if id_value in new_ids:
                                if id_value not in original_ids:
                                    write_to_cell(ws_output, output_row + i, column_index_from_string(col), id_value, Font(color="00FF00"))
                                else:
                                    write_to_cell(ws_output, output_row + i, column_index_from_string(col), id_value)
                            else:
                                write_to_cell(ws_output, output_row + i, column_index_from_string(col), id_value, Font(color="FF0000", strike=True))
                        
                        # Add merge range for columns B to F and M to N
                        merge_ranges_to_apply.append((start_output_row, column_index_from_string(col), output_row + row_span_A - 1, column_index_from_string(col)))
                    
                    # Handle column G and its dependent columns (H to L)
                    g_row = new_row
                    g_output_row = output_row
                    while g_row < new_row + row_span_A:
                        new_merged_range_G = get_merged_cell_range(ws_new, g_row, 'G')
                        if new_merged_range_G:
                            row_span_G = new_merged_range_G.max_row - new_merged_range_G.min_row + 1
                        else:
                            row_span_G = 1
                        
                        original_g_row = find_matching_row(ws_original, get_merged_cell_value(ws_new, g_row, 'G'), 'G')
                        
                        if original_g_row is not None:
                            original_merged_range_G = get_merged_cell_range(ws_original, original_g_row, 'G')
                            original_row_span_G = original_merged_range_G.max_row - original_merged_range_G.min_row + 1 if original_merged_range_G else 1
                            
                            # Handle G and H together
                            for col in ['G', 'H']:
                                new_ids = get_ids_in_range(ws_new, g_row, g_row + row_span_G - 1, col)
                                original_ids = get_ids_in_range(ws_original, original_g_row, original_g_row + original_row_span_G - 1, col)
                                
                                all_ids = list(set(new_ids + original_ids))
                                for i, id_value in enumerate(all_ids):
                                    if id_value in new_ids:
                                        if id_value not in original_ids:
                                            write_to_cell(ws_output, g_output_row + i, column_index_from_string(col), id_value, Font(color="00FF00"))
                                        else:
                                            write_to_cell(ws_output, g_output_row + i, column_index_from_string(col), id_value)
                                    else:
                                        write_to_cell(ws_output, g_output_row + i, column_index_from_string(col), id_value, Font(color="FF0000", strike=True))
                                
                                # Add merge range for columns G and H
                                merge_ranges_to_apply.append((g_output_row, column_index_from_string(col), g_output_row + row_span_G - 1, column_index_from_string(col)))
                            
                            # Handle I to L separately
                            for col in ['I', 'J', 'K', 'L']:
                                i_row = g_row
                                i_output_row = g_output_row
                                while i_row < g_row + row_span_G:
                                    new_merged_range_I = get_merged_cell_range(ws_new, i_row, col)
                                    if new_merged_range_I:
                                        row_span_I = new_merged_range_I.max_row - new_merged_range_I.min_row + 1
                                    else:
                                        row_span_I = 1
                                    
                                    original_i_row = find_matching_row(ws_original, get_merged_cell_value(ws_new, i_row, col), col)
                                    
                                    if original_i_row is not None:
                                        original_merged_range_I = get_merged_cell_range(ws_original, original_i_row, col)
                                        original_row_span_I = original_merged_range_I.max_row - original_merged_range_I.min_row + 1 if original_merged_range_I else 1
                                        
                                        new_ids = get_ids_in_range(ws_new, i_row, i_row + row_span_I - 1, col)
                                        original_ids = get_ids_in_range(ws_original, original_i_row, original_i_row + original_row_span_I - 1, col)
                                        
                                        all_ids = list(set(new_ids + original_ids))
                                        for i, id_value in enumerate(all_ids):
                                            if id_value in new_ids:
                                                if id_value not in original_ids:
                                                    write_to_cell(ws_output, i_output_row + i, column_index_from_string(col), id_value, Font(color="00FF00"))
                                                else:
                                                    write_to_cell(ws_output, i_output_row + i, column_index_from_string(col), id_value)
                                            else:
                                                write_to_cell(ws_output, i_output_row + i, column_index_from_string(col), id_value, Font(color="FF0000", strike=True))
                                    else:
                                        # New entry in column I-L, mark as green
                                        new_ids = get_ids_in_range(ws_new, i_row, i_row + row_span_I - 1, col)
                                        for i, id_value in enumerate(new_ids):
                                            write_to_cell(ws_output, i_output_row + i, column_index_from_string(col), id_value, Font(color="00FF00"))
                                    
                                    # Add merge range for columns I to L
                                    merge_ranges_to_apply.append((i_output_row, column_index_from_string(col), i_output_row + row_span_I - 1, column_index_from_string(col)))
                                    
                                    i_row += row_span_I
                                    i_output_row += row_span_I
                        else:
                            # New entry in column G, mark as green
                            for col in ['G', 'H']:
                                new_ids = get_ids_in_range(ws_new, g_row, g_row + row_span_G - 1, col)
                                for i, id_value in enumerate(new_ids):
                                    write_to_cell(ws_output, g_output_row + i, column_index_from_string(col), id_value, Font(color="00FF00"))
                                
                                # Add merge range for columns G and H
                                merge_ranges_to_apply.append((g_output_row, column_index_from_string(col), g_output_row + row_span_G - 1, column_index_from_string(col)))
                            
                            # Handle I to L separately for new entries
                            for col in ['I', 'J', 'K', 'L']:
                                i_row = g_row
                                i_output_row = g_output_row
                                while i_row < g_row + row_span_G:
                                    new_merged_range_I = get_merged_cell_range(ws_new, i_row, col)
                                    if new_merged_range_I:
                                        row_span_I = new_merged_range_I.max_row - new_merged_range_I.min_row + 1
                                    else:
                                        row_span_I = 1
                                    
                                    new_ids = get_ids_in_range(ws_new, i_row, i_row + row_span_I - 1, col)
                                    for i, id_value in enumerate(new_ids):
                                        write_to_cell(ws_output, i_output_row + i, column_index_from_string(col), id_value, Font(color="00FF00"))
                                    
                                    # Add merge range for columns I to L
                                    merge_ranges_to_apply.append((i_output_row, column_index_from_string(col), i_output_row + row_span_I - 1, column_index_from_string(col)))
                                    
                                    i_row += row_span_I
                                    i_output_row += row_span_I
                        
                        g_row += row_span_G
                        g_output_row += row_span_G
                    
                    output_row = g_output_row
                else:
                    # New entry in column A, mark entire range as green
                    for col in ['B', 'C', 'D', 'E', 'F', 'M', 'N']:
                        new_ids = get_ids_in_range(ws_new, new_row, new_row + row_span_A - 1, col)
                        for i, id_value in enumerate(new_ids):
                            write_to_cell(ws_output, output_row + i, column_index_from_string(col), id_value, Font(color="00FF00"))
                            
                        # Add merge range for columns B to F and M to N
                        merge_ranges_to_apply.append((start_output_row, column_index_from_string(col), output_row + row_span_A - 1, column_index_from_string(col)))
                    
                    # Handle columns G to L separately
                    g_row = new_row
                    g_output_row = output_row
                    while g_row < new_row + row_span_A or g_output_row < output_row + row_span_A:
                        new_merged_range_G = get_merged_cell_range(ws_new, g_row, 'G')
                        if new_merged_range_G:
                            row_span_G = new_merged_range_G.max_row - new_merged_range_G.min_row + 1
                        else:
                            row_span_G = 1
                        
                        for col in ['G', 'H']:
                            new_ids = get_ids_in_range(ws_new, g_row, g_row + row_span_G - 1, col)
                            for i, id_value in enumerate(new_ids):
                                write_to_cell(ws_output, g_output_row + i, column_index_from_string(col), id_value, Font(color="00FF00"))
                            
                            # Add merge range for columns G and H
                            merge_ranges_to_apply.append((g_output_row, column_index_from_string(col), g_output_row + row_span_G - 1, column_index_from_string(col)))
                        
                        # Handle I to L separately
                        for col in ['I', 'J', 'K', 'L']:
                            i_row = g_row
                            i_output_row = g_output_row
                            while i_row < g_row + row_span_G:
                                new_merged_range_I = get_merged_cell_range(ws_new, i_row, col)
                                if new_merged_range_I:
                                    row_span_I = new_merged_range_I.max_row - new_merged_range_I.min_row + 1
                                else:
                                    row_span_I = 1
                                
                                new_ids = get_ids_in_range(ws_new, i_row, i_row + row_span_I - 1, col)
                                for i, id_value in enumerate(new_ids):
                                    write_to_cell(ws_output, i_output_row + i, column_index_from_string(col), id_value, Font(color="00FF00"))
                                
                                # Add merge range for columns I to L
                                merge_ranges_to_apply.append((i_output_row, column_index_from_string(col), i_output_row + row_span_I - 1, column_index_from_string(col)))
                                
                                i_row += row_span_I
                                i_output_row += row_span_I
                        
                        g_row += row_span_G
                        g_output_row += row_span_G
                    
                    output_row = max(output_row + row_span_A, g_output_row)
                
                # Add merge range for the first column
                merge_ranges_to_apply.append((start_output_row, 1, output_row - 1, 1))
                
                new_row += row_span_A
            
            # Apply merge ranges after all content has been written
            for merge_range in merge_ranges_to_apply:
                ws_output.merge_cells(start_row=merge_range[0], start_column=merge_range[1],
                                      end_row=merge_range[2], end_column=merge_range[3])
            
            # Copy column widths from the new worksheet
            for col in range(1, ws_new.max_column + 1):
                col_letter = get_column_letter(col)
                ws_output.column_dimensions[col_letter].width = ws_new.column_dimensions[col_letter].width
    
    wb_output.remove(wb_output['Sheet'])
    wb_output.save(output_file)

def select_file(title):
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(title=title, filetypes=[("Excel files", "*.xlsx *.xls")])
    return file_path

def main():
    print("Excel Redline Comparison Tool")
    print("-----------------------------")
    
    original_file = select_file("Select the ORIGINAL Excel file")
    if not original_file:
        print("No original file selected. Exiting.")
        return

    new_file = select_file("Select the NEW Excel file")
    if not new_file:
        print("No new file selected. Exiting.")
        return

    output_file = filedialog.asksaveasfilename(
        title="Save the comparison result as",
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")]
    )
    if not output_file:
        print("No output file specified. Exiting.")
        return

    print("\nComparing files:")
    print(f"Original: {original_file}")
    print(f"New: {new_file}")
    print(f"Output: {output_file}")

    compare_excel_files(original_file, new_file, output_file)
    print(f"\nComparison complete. Results saved to {output_file}")

if __name__ == "__main__":
    main()