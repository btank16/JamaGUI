import openpyxl
import csv
from openpyxl.styles import Alignment
import tkinter as tk
from tkinter import messagebox, filedialog
import os
import shutil

def trace_view_export_formatter():
    """Main function to format Excel exports from Jama Connect's Trace View."""
    
    # Create root window and hide it
    root = tk.Tk()
    root.withdraw()
    
    # Ask for file path
    filepath = filedialog.askopenfilename(
        title="Select Excel file",
        filetypes=[
            ("Excel files", "*.xlsx"),
            ("All files", "*.*")
        ]
    )
    
    if not filepath:  # If user cancels file selection
        return
        
    # Verify file is xlsx
    if not filepath.lower().endswith('.xlsx'):
        messagebox.showerror("Error", "Please select a valid .xlsx file")
        return
    
    # Ask for output filename
    output_path = filedialog.asksaveasfilename(
        title="Save processed file as",
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        initialfile="processed_" + os.path.basename(filepath)
    )
    
    if not output_path:  # If user cancels save dialog
        return
    
    # Ask for header row
    header_row = tk.simpledialog.askinteger(
        "Header Row",
        "Enter row number for row containing table headers (default is 4)",
        initialvalue=4
    )
    if header_row is None:  # If user cancels
        return
        
    # Ask for harm column
    harm_column = tk.simpledialog.askstring(
        "Harm Input ID Column",
        "Enter the column letter for the Harm Input ID column (e.g., 'D')\nThis column and the next two will not be merged",
        initialvalue="D"
    )
    if harm_column is None:  # If user cancels
        return
        
    # Convert column letter to number (e.g., 'A'=1, 'B'=2, etc.)
    try:
        harm_col_num = openpyxl.utils.column_index_from_string(harm_column.upper())
    except:
        messagebox.showerror("Error", "Please enter a valid column letter")
        return
    
    # Copy the original file to the new location before processing
    shutil.copy2(filepath, output_path)
    
    merge_cells(output_path, header_row, harm_col_num)

def merge_cells(filepath, header_row, harm_col_num):
    """Merge cells in the Excel file based on column B values."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    # Find the maximum row and column
    max_row = ws.max_row
    max_col = ws.max_column
    
    # First, find merge ranges in column B
    merge_ranges = []
    current_value = None
    start_row = header_row + 1
    
    for row in range(header_row + 1, max_row + 1):
        cell_value = ws.cell(row=row, column=2).value
        
        if cell_value != current_value:
            if start_row and start_row < row - 1:
                merge_ranges.append((start_row, row - 1))
            current_value = cell_value
            start_row = row
    
    # Add the last range if necessary
    if start_row and start_row < max_row:
        merge_ranges.append((start_row, max_row))
    
    # Process each merge range independently
    for start_row, end_row in merge_ranges:
        # Merge column B
        if start_row < end_row:  # Only merge if there are multiple rows
            ws.merge_cells(f'B{start_row}:B{end_row}')
            merged_cell = ws.cell(row=start_row, column=2)
            merged_cell.alignment = Alignment(vertical='center')
        
        # For each column after B
        for col in range(3, max_col + 1):
            # Skip harm column and the next two columns
            if col == harm_col_num or col == harm_col_num + 1 or col == harm_col_num + 2:
                continue
                
            # Find sub-ranges within this range that have the same value
            sub_ranges = []
            sub_start = start_row
            current_value = ws.cell(row=start_row, column=col).value
            
            for row in range(start_row + 1, end_row + 1):
                cell_value = ws.cell(row=row, column=col).value
                
                if cell_value != current_value:
                    if sub_start < row - 1:  # Only add if there are multiple rows
                        sub_ranges.append((sub_start, row - 1))
                    sub_start = row
                    current_value = cell_value
            
            # Add the last sub-range if necessary
            if sub_start < end_row:
                sub_ranges.append((sub_start, end_row))
            
            # Merge each sub-range
            for sub_start, sub_end in sub_ranges:
                if sub_start < sub_end:  # Only merge if there are multiple rows
                    ws.merge_cells(
                        start_row=sub_start,
                        start_column=col,
                        end_row=sub_end,
                        end_column=col
                    )
                    merged_cell = ws.cell(row=sub_start, column=col)
                    merged_cell.alignment = Alignment(vertical='center')
    
    wb.save(filepath)

if __name__ == "__main__":
    trace_view_export_formatter()
