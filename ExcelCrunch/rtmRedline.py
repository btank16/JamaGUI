# Good for RTMs with 1 level Down

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter, column_index_from_string
import tkinter as tk
from tkinter import filedialog

def get_merged_cell_value(worksheet, row, col):
    """Get the value of a cell, taking into account if it's part of a merged range."""
    if isinstance(col, str):
        col = column_index_from_string(col)
    
    for merged_range in worksheet.merged_cells.ranges:
        if merged_range.min_row <= row <= merged_range.max_row and \
           merged_range.min_col <= col <= merged_range.max_col:
            return worksheet.cell(merged_range.min_row, merged_range.min_col).value
    return worksheet.cell(row, col).value

def get_merged_cell_range(worksheet, row, col):
    """Get the range of a merged cell if the cell is part of a merged range."""
    if isinstance(col, str):
        col = column_index_from_string(col)
    
    for merged_range in worksheet.merged_cells.ranges:
        if merged_range.min_row <= row <= merged_range.max_row and \
           merged_range.min_col <= col <= merged_range.max_col:
            return merged_range
    return None

def find_matching_row(worksheet, value, column):
    """Find the row in the given worksheet where the value in the specified column matches."""
    if isinstance(column, int):
        column = get_column_letter(column)
    
    for row in range(1, worksheet.max_row + 1):
        if get_merged_cell_value(worksheet, row, column) == value:
            return row
    return None

def get_ids_in_merged_range(worksheet, merged_range, column):
    """Get all IDs within a merged range for a specific column."""
    ids = []
    for row in range(merged_range.min_row, merged_range.max_row + 1):
        value = worksheet.cell(row, column).value
        if value is not None:
            ids.append(value)
    return ids

def write_to_cell(worksheet, row, col, value, font=None):
    """Write to a cell, handling merged cells by writing to the top-left cell."""
    cell = worksheet.cell(row, col)
    cell.value = value
    if font:
        cell.font = font

def compare_excel_files(original_file, new_file, output_file):
    wb_original = openpyxl.load_workbook(original_file)
    wb_new = openpyxl.load_workbook(new_file)
    wb_output = openpyxl.Workbook()
    
    for sheet_name in wb_new.sheetnames:
        if sheet_name in wb_original.sheetnames:
            ws_original = wb_original[sheet_name]
            ws_new = wb_new[sheet_name]
            ws_output = wb_output.create_sheet(sheet_name)
            
            output_row = 1
            processed_first_column_ids = set()
            merge_ranges_to_apply = []
            
            for col in range(1, ws_new.max_column):
                col_letter = get_column_letter(col)
                next_col_letter = get_column_letter(col + 1)
                
                for new_row in range(1, ws_new.max_row + 1):
                    key_value = get_merged_cell_value(ws_new, new_row, col_letter)
                    
                    if col == 1:
                        if key_value in processed_first_column_ids:
                            continue
                        processed_first_column_ids.add(key_value)
                    
                    original_row = find_matching_row(ws_original, key_value, col_letter)
                    
                    # Write the key value to the output
                    write_to_cell(ws_output, output_row, col, key_value)
                    
                    new_merged_range = get_merged_cell_range(ws_new, new_row, col_letter)
                    if new_merged_range:
                        row_span = new_merged_range.max_row - new_merged_range.min_row + 1
                    else:
                        row_span = 1
                    
                    start_output_row = output_row
                    
                    if original_row is not None:
                        original_merged_range = get_merged_cell_range(ws_original, original_row, col_letter)
                        
                        new_ids = get_ids_in_merged_range(ws_new, new_merged_range, col + 1) if new_merged_range else [ws_new.cell(new_row, col + 1).value]
                        original_ids = get_ids_in_merged_range(ws_original, original_merged_range, col + 1) if original_merged_range else [ws_original.cell(original_row, col + 1).value]
                        
                        all_ids = list(set(new_ids + original_ids))
                        for i, id_value in enumerate(all_ids):
                            if id_value in new_ids:
                                if id_value not in original_ids:
                                    write_to_cell(ws_output, output_row + i, col + 1, id_value, Font(color="00FF00"))
                                else:
                                    write_to_cell(ws_output, output_row + i, col + 1, id_value)
                            else:
                                write_to_cell(ws_output, output_row + i, col + 1, id_value, Font(color="FF0000", strike=True))
                        
                        row_span = max(row_span, len(all_ids))
                    else:
                        # New entry, mark entire merged range (or single cell) as green
                        if new_merged_range:
                            new_ids = get_ids_in_merged_range(ws_new, new_merged_range, col + 1)
                            for i, id_value in enumerate(new_ids):
                                write_to_cell(ws_output, output_row + i, col, key_value, Font(color="00FF00"))
                                write_to_cell(ws_output, output_row + i, col + 1, id_value, Font(color="00FF00"))
                            row_span = max(row_span, len(new_ids))
                        else:
                            write_to_cell(ws_output, output_row, col, key_value, Font(color="00FF00"))
                            new_id = ws_new.cell(new_row, col + 1).value
                            if new_id is not None:
                                write_to_cell(ws_output, output_row, col + 1, new_id, Font(color="00FF00"))
                    
                    # Add merge range if necessary
                    if row_span > 1:
                        merge_range = (start_output_row, col, output_row + row_span - 1, col)
                        merge_ranges_to_apply.append(merge_range)
                    
                    output_row += row_span
            
            # Apply merge ranges after all content has been written
            for merge_range in merge_ranges_to_apply:
                ws_output.merge_cells(start_row=merge_range[0], start_column=merge_range[1],
                                      end_row=merge_range[2], end_column=merge_range[3])
            
            for col in range(1, ws_new.max_column + 1):
                col_letter = get_column_letter(col)
                ws_output.column_dimensions[col_letter].width = ws_new.column_dimensions[col_letter].width
    
    wb_output.remove(wb_output['Sheet'])
    wb_output.save(output_file)

def select_file(title):
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(title=title, filetypes=[("Excel files", "*.xlsx *.xls")])
    return file_path

def main():
    print("Excel Redline Comparison Tool")
    print("-----------------------------")
    
    original_file = select_file("Select the ORIGINAL Excel file")
    if not original_file:
        print("No original file selected. Exiting.")
        return

    new_file = select_file("Select the NEW Excel file")
    if not new_file:
        print("No new file selected. Exiting.")
        return

    output_file = filedialog.asksaveasfilename(
        title="Save the comparison result as",
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")]
    )
    if not output_file:
        print("No output file specified. Exiting.")
        return

    print("\nComparing files:")
    print(f"Original: {original_file}")
    print(f"New: {new_file}")
    print(f"Output: {output_file}")

    compare_excel_files(original_file, new_file, output_file)
    print(f"\nComparison complete. Results saved to {output_file}")

if __name__ == "__main__":
    main()